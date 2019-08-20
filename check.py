#as of 20AUG2019


from __future__ import print_function
import os
import sys
import random
from datetime import *
from openpyxl import *
import math
import csv


#######Global VARS
sheet1 = 'RCC Operations 2019'
sheet2 = 'RCC C4IM Services 2019'
sheet3 = 'RCC Administrative 2019'

#vars for telling which rows are meaningful might need to find better way to do this
appOpsRow = range(7, 236,1)
appC4Row = range(7, 191,1)
appADRow = range(7, 230,1)


###################


#working on these functions

#returns a list of all tasks
def createTaskList(excel):
    taskList = []
    try:
        wb = load_workbook(filename=excel)
    except:
        print("Error loading sheet")
    ws = wb.get_sheet_by_name(sheet1)
    
    for x in appOpsRow:
        taskList.append(ws['G'+str(x)].value)
    
    ws = wb.get_sheet_by_name(sheet2)
    for x in appC4Row:
        taskList.append(ws['G'+str(x)].value)
    
    ws = wb.get_sheet_by_name(sheet3)
    for x in appADRow:
        taskList.append(ws['H'+str(x)].value)
    return taskList

#sum members of a list that might have nonetypes in it
def wierdSum(wList):
    wSum = 0
    for x in wList:
        if type(x) == long or type(x) == int or type(x) == float:
            wSum += x
    return wSum
    
    
#returns a number of submitted hours from an individual (list of dictionaries of their submissions)
def submitHours(dictDayList):
    hoursTot = 0
    for day in dictDayList:
        for task in day:
            hoursTot += wierdSum(day[task])
    return hoursTot


#returns list of tasks worked for an individual
def submitTasks(dictDayList):
    workedTask = []
    for day in dictDayList:
        for task in day:
            if wierdSum(day[task]) > 0:
                workedTask.append(task)
    return workedTask


#returns tasks not worked by all





#finished functions
def getPax(dir):
    paxholder = []
    for root, dirs, files in os.walk(dir):
        for x in dirs:
			paxholder.append(x)
    return paxholder


def combineDicts(dictList):
    newDict = {}
    for Dict in dictList:
        for keys in Dict:
            newDict[keys] = Dict[keys]
    return newDict


def getDayData(excel):
	try :
		wb = load_workbook(filename = excel)
	except:
		print("Error loading sheet")
	#prep containers and sheet data
	Ops = {}
	C4IM = {}
	Admin = {}

	#retrieve data from Ops sheet
	#construct list containing rows in sheet to get data from

	ws = wb.get_sheet_by_name(sheet1)
	for x in appOpsRow:
		currkey = ws['G'+str(x)].value
		currdata = [ ws['Z'+ str(x)].value , ws['AD'+ str(x)].value, ws['AH'+ str(x)].value ]

		Ops[currkey] = currdata

	#retrieve data from C4IM sheet
	ws = wb.get_sheet_by_name(sheet2)
	for x in appC4Row:
		currkey = ws['G'+str(x)].value
		currdata = [ ws['Z'+ str(x)].value , ws['AH'+ str(x)].value , ws['AN'+ str(x)].value ]

		C4IM[currkey] = currdata
	#print (C4IM['700.4.4.1  Operate and maintain web monitoring and filtering systems IAW applicable laws and regulations'])
	#retrieve data from Admin sheet
	ws = wb.get_sheet_by_name(sheet3)
	for x in appADRow:
		currkey = ws['H'+str(x)].value
		currdata = [ ws['N'+ str(x)].value , ws['T'+ str(x)].value , ws['V'+ str(x)].value ]

		Admin[currkey] = currdata
	#print (Admin['Monitor & submit award recommendations; Initiate & route non-competitive actions via DCPDS and via AUTO-NOA for competitive actions to G1'])
	retDict = combineDicts([Ops, C4IM, Admin])
	return retDict


def getDaysinFolder(folder):
    dayList=[]
    currfolder = ('./Submissions/'+folder+'/')
    for root,dirs, files in os.walk(currfolder):
        for x in files:
            dayList.append(getDayData('./Submissions/'+folder+'/'+x))
    return dayList


def main():
    TList = createTaskList('RCC_Manpowerstudy-AllMaster.xlsx')
    
    #csv of hours submitted by employee
    with open('Individual_SubmissionStat.csv', 'w') as IndivCSV:
        indiWriter = csv.writer(IndivCSV, delimiter = ',')
        people = getPax('./Submissions')
        
        for person in people:
            daySheets = getDaysinFolder(person)
            print(person, submitHours(daySheets))
            indiWriter.writerow([person, submitHours(daySheets)])
    
    
        #taskHolder = submitTasks(daySheets)
        #for x in taskHolder:
        #    print (x)
       
        

    
    



if __name__ == '__main__':
    print ("***running***")
    main()